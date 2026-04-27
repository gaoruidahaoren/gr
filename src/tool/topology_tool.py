from collections import defaultdict
from typing import Any

from fastapi import HTTPException, UploadFile

from src.model.settings import FIXED_TOPOLOGY
from src.utli.common import name_hit, normalize_meter_date, normalize_name, parse_date, safe_float


def infer_hierarchy(nodes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_id = {n["id"]: n for n in nodes}
    for n in nodes:
        if n.get("parent_id"):
            continue
        meter_id = n["id"]
        import re
        m_l3 = re.match(r"^L3-(\d+)-\d+$", meter_id)
        if m_l3:
            parent_id = f"L2-{m_l3.group(1)}"
            if parent_id in by_id:
                n["parent_id"] = parent_id
                if n["level"] < 3:
                    n["level"] = 3
                continue
        m_l2 = re.match(r"^L2-(\d+)$", meter_id)
        if m_l2 and "L1" in by_id:
            n["parent_id"] = "L1"
            if n["level"] < 2:
                n["level"] = 2

    roots = [n for n in nodes if "总表" in n["name"]]
    if len(roots) == 1:
        root = roots[0]
        if root["level"] < 2:
            root["level"] = 2
        for n in nodes:
            if n["id"] == root["id"]:
                continue
            if not n.get("parent_id"):
                n["parent_id"] = root["id"]
                if n["level"] < 3:
                    n["level"] = 3

    return nodes


def _topology_children_map() -> dict[str, list[str]]:
    child_map: dict[str, list[str]] = defaultdict(list)
    for cid, cfg in FIXED_TOPOLOGY.items():
        parent = cfg["parent"]
        if parent:
            child_map[parent].append(cid)
    return child_map


def _find_canonical_id(node: dict[str, Any], used_ids: set[str]) -> str | None:
    node_id = node["id"]
    if node_id in FIXED_TOPOLOGY and node_id not in used_ids:
        cfg = FIXED_TOPOLOGY[node_id]
        if name_hit(node["name"], cfg["aliases"]) or normalize_name(node["name"]) in ("", normalize_name(node_id)):
            return node_id

    candidates: list[tuple[int, str]] = []
    for cid, cfg in FIXED_TOPOLOGY.items():
        if cid in used_ids:
            continue
        aliases = cfg["aliases"]
        if name_hit(node["name"], aliases):
            score = max(len(normalize_name(a)) for a in aliases)
            candidates.append((score, cid))

    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


def _latest_reading_value(node: dict[str, Any]) -> float:
    readings = node.get("readings", [])
    if not readings:
        return 0.0
    return safe_float(readings[-1].get("value", 0.0))


def ensure_office_total_present(nodes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    present_ids = {n["id"] for n in nodes}
    office_subs = {"L3-6-1", "L3-6-2", "L3-6-3"}
    has_office_sub = bool(present_ids & office_subs)
    if not has_office_sub or "L2-6" in present_ids:
        return nodes

    candidates: list[dict[str, Any]] = []
    for n in nodes:
        if n["id"] == "L1":
            candidates.append(n)
    for n in nodes:
        if not n.get("is_fixed_topology"):
            candidates.append(n)

    if not candidates:
        return nodes

    candidates.sort(key=_latest_reading_value, reverse=True)
    cand = candidates[0]
    cfg = FIXED_TOPOLOGY["L2-6"]
    cand["raw_name"] = cand.get("name", "")
    cand["id"] = "L2-6"
    cand["name"] = cfg["name"]
    cand["parent_id"] = cfg["parent"]
    cand["level"] = cfg["level"]
    cand["multiplier"] = float(cfg["multiplier"])
    cand["meter_type"] = "total"
    cand["is_fixed_topology"] = True
    return nodes


def fill_missing_office_children(nodes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    present_ids = {n["id"] for n in nodes}
    if "L2-6" not in present_ids:
        return nodes

    required_children = ["L3-6-1", "L3-6-2", "L3-6-3"]
    missing = [cid for cid in required_children if cid not in present_ids]
    if not missing:
        return nodes

    extras = [n for n in nodes if not n.get("is_fixed_topology")]
    if not extras:
        return nodes

    extras.sort(key=_latest_reading_value)
    ordered_missing = sorted(missing, key=lambda x: 0 if x == "L3-6-1" else 1)

    for cid in ordered_missing:
        if not extras:
            break
        candidate = extras.pop(0)
        cfg = FIXED_TOPOLOGY[cid]
        candidate["raw_name"] = candidate.get("name", "")
        candidate["id"] = cid
        candidate["name"] = cfg["name"]
        candidate["parent_id"] = cfg["parent"]
        candidate["level"] = cfg["level"]
        candidate["multiplier"] = float(cfg["multiplier"])
        candidate["meter_type"] = "sub"
        candidate["is_fixed_topology"] = True

    return nodes


def bind_fixed_topology(nodes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    used_ids: set[str] = set()
    bound_nodes: list[dict[str, Any]] = []

    for n in nodes:
        canonical_id = _find_canonical_id(n, used_ids)
        if canonical_id:
            cfg = FIXED_TOPOLOGY[canonical_id]
            n["raw_name"] = n["name"]
            n["id"] = canonical_id
            n["name"] = cfg["name"]
            n["parent_id"] = cfg["parent"]
            n["level"] = cfg["level"]
            n["multiplier"] = float(cfg["multiplier"])
            n["meter_type"] = "sub"
            n["is_fixed_topology"] = True
            used_ids.add(canonical_id)
        else:
            n["is_fixed_topology"] = False
        bound_nodes.append(n)

    child_map = _topology_children_map()
    for n in bound_nodes:
        if n["id"] in child_map:
            n["meter_type"] = "total"

    bound_nodes = ensure_office_total_present(bound_nodes)
    bound_nodes = fill_missing_office_children(bound_nodes)
    return bound_nodes


def filter_scope_nodes(nodes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ids = {n["id"] for n in nodes}
    office_ids = {"L2-6", "L3-6-1", "L3-6-2", "L3-6-3"}
    other_branch_ids = {"L2-1", "L2-2", "L2-3", "L2-4", "L2-5", "L3-1-1", "L3-1-2"}

    has_office = bool(ids & office_ids)
    has_other = bool(ids & other_branch_ids)

    if has_office and not has_other:
        return [n for n in nodes if n["id"] in office_ids]
    return nodes


def normalize_payload(payload: dict[str, Any], meter_date: str | None = None) -> dict[str, Any]:
    nodes = payload.get("nodes", [])
    if not isinstance(nodes, list):
        raise HTTPException(status_code=400, detail="nodes 必须是数组")

    chosen_date = normalize_meter_date(meter_date)
    normalized_nodes: list[dict[str, Any]] = []

    for idx, n in enumerate(nodes):
        if not isinstance(n, dict):
            continue
        meter_id = str(n.get("id") or f"N-{idx + 1}")
        name = str(n.get("name") or meter_id)
        parent_id = n.get("parent_id")
        parent_id = str(parent_id) if parent_id is not None else None
        level = int(safe_float(n.get("level"), 1))
        multiplier = safe_float(n.get("multiplier"), 1.0) or 1.0

        readings: list[dict[str, Any]] = []
        rs = n.get("readings", [])
        if isinstance(rs, list):
            for r in rs:
                if not isinstance(r, dict):
                    continue
                value = r.get("value")
                if value is None:
                    continue
                readings.append({"date": chosen_date, "value": safe_float(value)})
        readings.sort(key=lambda x: x["date"])

        normalized_nodes.append({
            "id": meter_id,
            "name": name,
            "meter_type": str(n.get("meter_type") or "").strip().lower(),
            "parent_id": parent_id,
            "level": level,
            "multiplier": multiplier,
            "readings": readings,
        })

    if not normalized_nodes:
        raise HTTPException(status_code=400, detail="模型未识别出有效电表节点")

    normalized_nodes = infer_hierarchy(normalized_nodes)
    normalized_nodes = bind_fixed_topology(normalized_nodes)
    normalized_nodes = filter_scope_nodes(normalized_nodes)

    return {
        "time_range": payload.get("time_range", ""),
        "company": payload.get("company", ""),
        "nodes": normalized_nodes,
    }


def load_previous_excel(file: UploadFile) -> dict[str, Any]:
    content = file.file.read()
    from io import BytesIO
    from openpyxl import load_workbook

    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"历史Excel读取失败: {exc}") from exc

    if "识别原始数据" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="历史Excel缺少sheet: 识别原始数据")

    ws = wb["识别原始数据"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        raise HTTPException(status_code=400, detail="历史Excel格式不正确")

    company = str(rows[0][1] or "")
    time_range = str(rows[1][1] or "")
    headers_raw = [str(x) if x is not None else "" for x in rows[3]]
    headers = [h.strip() for h in headers_raw]

    def _find_idx(*names: str) -> int | None:
        for i, h in enumerate(headers):
            if h in names:
                return i
        return None

    idx_id = _find_idx("ID", "编号")
    idx_name = _find_idx("名称", "用电部位", "电表名称")
    idx_type = _find_idx("电表类型")
    idx_parent = _find_idx("父ID")
    idx_level = _find_idx("层级")
    idx_multi = _find_idx("倍率")

    if idx_id is None or idx_name is None:
        raise HTTPException(status_code=400, detail="历史Excel缺少 ID/名称 列")

    known = {x for x in [idx_id, idx_name, idx_type, idx_parent, idx_level, idx_multi] if x is not None}
    date_indexes = [i for i, h in enumerate(headers) if i not in known and h]
    if not date_indexes:
        raise HTTPException(status_code=400, detail="历史Excel缺少日期列")

    nodes: list[dict[str, Any]] = []
    for r in rows[4:]:
        if not r:
            continue
        meter_raw = r[idx_id] if idx_id < len(r) else None
        if meter_raw in (None, ""):
            continue
        meter_id = str(meter_raw)
        name_raw = r[idx_name] if idx_name < len(r) else None
        name = str(name_raw or meter_id)
        meter_type = str((r[idx_type] if idx_type is not None and idx_type < len(r) else "") or "")
        parent_raw = r[idx_parent] if idx_parent is not None and idx_parent < len(r) else None
        parent_id = str(parent_raw) if parent_raw not in (None, "") else None
        level_raw = r[idx_level] if idx_level is not None and idx_level < len(r) else 1
        level = int(safe_float(level_raw, 1))
        multi_raw = r[idx_multi] if idx_multi is not None and idx_multi < len(r) else 1.0
        multiplier = safe_float(multi_raw, 1.0) or 1.0

        readings: list[dict[str, Any]] = []
        for i in date_indexes:
            d = headers[i]
            val = r[i] if i < len(r) else None
            if val in (None, ""):
                continue
            readings.append({"date": str(d), "value": safe_float(val)})

        nodes.append({
            "id": meter_id,
            "name": name,
            "meter_type": meter_type,
            "parent_id": parent_id,
            "level": level,
            "multiplier": multiplier,
            "readings": readings,
        })

    if not nodes:
        raise HTTPException(status_code=400, detail="历史Excel中没有有效电表记录")

    return {"company": company, "time_range": time_range, "nodes": nodes}


def merge_previous_and_today(previous: dict[str, Any], today: dict[str, Any]) -> dict[str, Any]:
    node_map: dict[str, dict[str, Any]] = {}
    for src in previous.get("nodes", []):
        node_map[src["id"]] = {
            "id": src["id"],
            "name": src["name"],
            "meter_type": src.get("meter_type", ""),
            "parent_id": src.get("parent_id"),
            "level": src.get("level", 1),
            "multiplier": src.get("multiplier", 1.0),
            "readings": list(src.get("readings", [])),
        }

    for src in today.get("nodes", []):
        target = node_map.get(src["id"])
        if target is None:
            target = {
                "id": src["id"],
                "name": src["name"],
                "meter_type": src.get("meter_type", ""),
                "parent_id": src.get("parent_id"),
                "level": src.get("level", 1),
                "multiplier": src.get("multiplier", 1.0),
                "readings": [],
            }
            node_map[src["id"]] = target

        target["name"] = src["name"]
        target["meter_type"] = src.get("meter_type", target.get("meter_type", ""))
        target["parent_id"] = src.get("parent_id")
        target["level"] = src.get("level", target.get("level", 1))
        target["multiplier"] = src.get("multiplier", target.get("multiplier", 1.0))

        prev_by_date = {r["date"]: r for r in target["readings"]}
        for r in src.get("readings", []):
            prev_by_date[r["date"]] = {"date": r["date"], "value": safe_float(r["value"])}
        merged = list(prev_by_date.values())
        try:
            merged.sort(key=lambda x: parse_date(x["date"]))
        except ValueError:
            merged.sort(key=lambda x: x["date"])
        target["readings"] = merged

    merged_nodes = list(node_map.values())
    merged_nodes = bind_fixed_topology(merged_nodes)
    merged_nodes = filter_scope_nodes(merged_nodes)

    return {
        "company": today.get("company") or previous.get("company", ""),
        "time_range": today.get("time_range") or previous.get("time_range", ""),
        "nodes": merged_nodes,
    }


def topology_children_map() -> dict[str, list[str]]:
    return _topology_children_map()
