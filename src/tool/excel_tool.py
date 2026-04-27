from collections import defaultdict
from tempfile import NamedTemporaryFile
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from src.tool.topology_tool import topology_children_map
from src.utli.common import safe_float


def compute_usage(data: dict[str, Any]) -> tuple[list[str], list[dict[str, Any]], list[dict[str, Any]]]:
    nodes = data["nodes"]
    node_by_id = {n["id"]: n for n in nodes}
    children_map: dict[str, list[str]] = defaultdict(list)
    for n in nodes:
        pid = n["parent_id"]
        if pid and pid in node_by_id:
            children_map[pid].append(n["id"])

    all_dates = set()
    for n in nodes:
        for r in n["readings"]:
            all_dates.add(r["date"])
    dates = sorted(all_dates)
    if not dates:
        raise HTTPException(status_code=400, detail="未识别到有效日期读数")

    use_diff_mode = len(dates) >= 2
    metric_labels = [f"{dates[i]}->{dates[i + 1]}" for i in range(len(dates) - 1)] if use_diff_mode else dates[:]

    usage_rows: list[dict[str, Any]] = []
    usage_map: dict[str, list[float]] = {}

    for n in nodes:
        reading_map = {r["date"]: safe_float(r["value"]) for r in n["readings"]}
        if use_diff_mode:
            usage_list: list[float] = []
            for i in range(len(dates) - 1):
                d0, d1 = dates[i], dates[i + 1]
                if d0 in reading_map and d1 in reading_map:
                    usage_list.append(round((reading_map[d1] - reading_map[d0]) * n["multiplier"], 2))
                else:
                    usage_list.append(0.0)
        else:
            usage_list = [round(reading_map.get(d, 0.0) * n["multiplier"], 2) for d in dates]

        usage_map[n["id"]] = usage_list
        usage_rows.append({
            "id": n["id"],
            "name": n["name"],
            "meter_type": n.get("meter_type", ""),
            "level": n["level"],
            "parent_id": n["parent_id"],
            "multiplier": n["multiplier"],
            "daily_usage": usage_list,
        })

    balance_rows: list[dict[str, Any]] = []
    expected_children = topology_children_map()
    present_ids = {n["id"] for n in nodes}

    for n in nodes:
        expected_child_ids = expected_children.get(n["id"], [])
        present_child_ids = [cid for cid in expected_child_ids if cid in present_ids]
        if not present_child_ids:
            present_child_ids = children_map.get(n["id"], [])
        if not present_child_ids:
            continue

        node_usage = usage_map[n["id"]]
        child_sum = [0.0] * len(node_usage)
        for cid in present_child_ids:
            cu = usage_map[cid]
            for i, v in enumerate(cu):
                child_sum[i] += v

        missing = [cid for cid in expected_child_ids if cid not in present_child_ids]
        diffs = [round(node_usage[i] - child_sum[i], 2) for i in range(len(node_usage))]
        balance_rows.append({
            "id": n["id"],
            "name": n["name"],
            "children": ", ".join(present_child_ids),
            "node_usage": [round(v, 2) for v in node_usage],
            "children_sum": [round(v, 2) for v in child_sum],
            "diff": diffs,
            "expected_children": ", ".join(expected_child_ids),
            "missing_children": ", ".join(missing),
            "child_scope": "完整分支" if not missing and expected_child_ids else "部分分支",
        })

    return metric_labels, usage_rows, balance_rows


def build_excel(data: dict[str, Any], dates: list[str], usage_rows: list[dict[str, Any]], balance_rows: list[dict[str, Any]]) -> str:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "识别原始数据"

    ws1.append(["公司", data.get("company", "")])
    ws1.append(["时间范围", data.get("time_range", "")])
    ws1.append([])

    all_dates = sorted({r["date"] for n in data["nodes"] for r in n["readings"]})
    headers = ["ID", "名称", "电表类型", "父ID", "层级", "倍率"] + all_dates
    ws1.append(headers)

    for n in data["nodes"]:
        reading_map = {r["date"]: r["value"] for r in n["readings"]}
        row = [n["id"], n["name"], n.get("meter_type", ""), n["parent_id"] or "", n["level"], n["multiplier"]]
        row.extend([reading_map.get(d, "") for d in all_dates])
        ws1.append(row)

    ws2 = wb.create_sheet("当日用电量")
    ws2.append(["ID", "名称", "电表类型", "层级", "父ID", "倍率"] + dates)
    for r in usage_rows:
        ws2.append([r["id"], r["name"], r.get("meter_type", ""), r["level"], r["parent_id"] or "", r["multiplier"]] + r["daily_usage"])

    ws3 = wb.create_sheet("差值校核")
    ws3.append(["父节点ID", "父节点", "已上传子节点", "预期子节点", "缺失子节点", "分支范围", "指标"] + dates)
    for b in balance_rows:
        prefix = [b["id"], b["name"], b["children"], b.get("expected_children", ""), b.get("missing_children", ""), b.get("child_scope", "")]
        ws3.append(prefix + ["父表用电量"] + b["node_usage"])
        ws3.append(prefix + ["子表合计"] + b["children_sum"])
        ws3.append(prefix + ["差值(父-子)"] + b["diff"])

    for ws in [ws1, ws2, ws3]:
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 16
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    return tmp.name
